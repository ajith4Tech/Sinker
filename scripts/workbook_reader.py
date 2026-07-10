"""Reads a saved .xlsx back with openpyxl into a plain-JSON "workbook
model" — cell values *and* formatting (merges, borders, fills, fonts,
alignment, row heights, column widths, freeze panes).

This is PART 2/3 of the fix: the preview must render the actual generated
workbook, never parser JSON. This module is the only place that produces the
model the frontend renders, and it's built by opening the .xlsx file
scripts/parser.py just saved — the same bytes served by the download route.

Styles are de-duplicated into a `styles` table (cells reference a styleId)
since most cells in a large sheet share one of a handful of distinct looks —
this keeps the JSON payload from repeating identical style objects per cell.
"""

from __future__ import annotations

from datetime import date, datetime, time

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

DEFAULT_COLUMN_WIDTH = 8.43
DEFAULT_ROW_HEIGHT = 15.0


def read_workbook(path: str) -> dict:
    workbook = load_workbook(path)
    sheet = workbook.active

    max_row = sheet.max_row
    max_col = sheet.max_column

    style_index: dict[tuple, int] = {}
    styles: list[dict] = []

    def style_id_for(cell) -> int:
        key = _style_key(cell)
        if key not in style_index:
            style_index[key] = len(styles)
            styles.append(_style_dict(cell))
        return style_index[key]

    rows = []
    for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        rows.append([
            {"value": _serialize_value(cell.value), "styleId": style_id_for(cell)}
            for cell in row
        ])

    merges = [
        {
            "minRow": r.min_row, "maxRow": r.max_row,
            "minCol": r.min_col, "maxCol": r.max_col,
        }
        for r in sheet.merged_cells.ranges
    ]

    column_widths = [
        sheet.column_dimensions[get_column_letter(c)].width or DEFAULT_COLUMN_WIDTH
        for c in range(1, max_col + 1)
    ]
    row_heights = [
        sheet.row_dimensions[r].height or DEFAULT_ROW_HEIGHT
        for r in range(1, max_row + 1)
    ]

    freeze = sheet.freeze_panes  # e.g. "A4" or None

    return {
        "sheetName": sheet.title,
        "maxRow": max_row,
        "maxCol": max_col,
        "columnWidths": column_widths,
        "rowHeights": row_heights,
        "merges": merges,
        "freezePanes": freeze,
        "styles": styles,
        "rows": rows,
    }


def _serialize_value(value):
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value


def _color_rgb(color) -> str | None:
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str):
        return rgb
    return None


def _border_side(side) -> dict | None:
    if side is None or side.style is None:
        return None
    return {"style": side.style, "color": _color_rgb(side.color)}


def _style_key(cell) -> tuple:
    font = cell.font
    fill = cell.fill
    border = cell.border
    align = cell.alignment
    return (
        bool(font.bold), bool(font.italic), font.size, _color_rgb(font.color), font.name,
        fill.fill_type, _color_rgb(fill.fgColor) if fill.fill_type and fill.fill_type != "none" else None,
        border.top.style if border.top else None, _color_rgb(border.top.color) if border.top else None,
        border.bottom.style if border.bottom else None, _color_rgb(border.bottom.color) if border.bottom else None,
        border.left.style if border.left else None, _color_rgb(border.left.color) if border.left else None,
        border.right.style if border.right else None, _color_rgb(border.right.color) if border.right else None,
        align.horizontal, align.vertical, bool(align.wrap_text),
        cell.number_format,
    )


def _style_dict(cell) -> dict:
    font = cell.font
    fill = cell.fill
    border = cell.border
    align = cell.alignment
    return {
        "font": {
            "bold": bool(font.bold),
            "italic": bool(font.italic),
            "size": font.size,
            "color": _color_rgb(font.color),
            "name": font.name,
        },
        "fill": {
            "color": _color_rgb(fill.fgColor) if fill.fill_type and fill.fill_type != "none" else None,
        },
        "border": {
            "top": _border_side(border.top),
            "bottom": _border_side(border.bottom),
            "left": _border_side(border.left),
            "right": _border_side(border.right),
        },
        "alignment": {
            "horizontal": align.horizontal,
            "vertical": align.vertical,
            "wrapText": bool(align.wrap_text),
        },
        "numberFormat": cell.number_format,
    }
