"""ExcelWriter: load a template workbook, append new rows, save a copy.

Knows nothing about PDF parsing — it only ever sees LineRecord objects and a
de-duplication key string. Loads the template with openpyxl (which preserves
existing formatting, merged cells, formulas, fonts, borders, and column
widths as long as we only ever write to previously-untouched data cells) and
writes to a *new* output path; the template itself is never modified.

Performance: the workbook is loaded once, the existing-row HashSet is built
once by a single top-to-bottom scan, and the workbook is saved once after
every row has been appended in memory.
"""

from __future__ import annotations

from datetime import date, datetime, time

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from models import LineRecord
from template_schema import (
    ATTRS,
    COLUMNS,
    FIRST_DATA_ROW,
    INVOICE_COL,
    ITEM_COL,
    NUM_COLUMNS,
    NUMERIC_ATTRS,
    SHIPPING_BILL_COL,
)

_RIGHT_ALIGN = Alignment(horizontal="right")


def _coerce(attr: str, value):
    """Numeric-column values are written as real numbers (not text) so
    Excel right-aligns them and its own number semantics (sorting, SUM
    formulas) work — never guessed for non-numeric columns, and left as the
    original string if it doesn't parse cleanly (never silently dropped)."""
    if attr not in NUMERIC_ATTRS or value is None:
        return value
    try:
        as_float = float(value)
    except (TypeError, ValueError):
        return value
    return int(as_float) if as_float.is_integer() else as_float


class ExcelWriter:
    def __init__(self, template_path: str):
        self._workbook = load_workbook(template_path)
        self._sheet = self._workbook.active
        self._seen_keys: set[str] = set()
        self._new_rows: set[int] = set()
        self._next_row = self._scan_existing_rows()

    def _scan_existing_rows(self) -> int:
        """Single pass over existing data rows: builds the de-duplication
        HashSet and finds the true last data row (ignoring trailing
        formatted-but-empty rows some templates ship with)."""
        last_data_row = FIRST_DATA_ROW - 1

        for row in range(FIRST_DATA_ROW, self._sheet.max_row + 1):
            sb = self._sheet.cell(row=row, column=SHIPPING_BILL_COL).value
            invoice = self._sheet.cell(row=row, column=INVOICE_COL).value
            item = self._sheet.cell(row=row, column=ITEM_COL).value

            if sb is None and invoice is None and item is None:
                continue

            last_data_row = row
            self._seen_keys.add(f"{sb}|{invoice}|{item}")

        return last_data_row + 1

    def has_key(self, key: str) -> bool:
        return key in self._seen_keys

    def append_row(self, key: str, record: LineRecord) -> dict[str, str]:
        """Appends one LineRecord. Caller must have already checked
        has_key(key). Returns {field_header: "A1"-style cell address} for
        every non-spacer column, so callers can back-fill workbook_cell on
        debug_log entries."""
        row = self._next_row
        cell_map: dict[str, str] = {}
        values = record.to_dict()

        for col_index, (header, attr) in enumerate(zip(COLUMNS, ATTRS), start=1):
            if attr is None:
                continue
            cell = self._sheet.cell(row=row, column=col_index, value=_coerce(attr, values.get(attr)))
            if attr in NUMERIC_ATTRS:
                cell.alignment = _RIGHT_ALIGN
            cell_map[header] = f"{get_column_letter(col_index)}{row}"

        self._seen_keys.add(key)
        self._new_rows.add(row)
        self._next_row += 1
        return cell_map

    def save(self, output_path: str) -> None:
        self._workbook.save(output_path)

    def to_preview(self) -> dict:
        """Converts every data row currently in the worksheet (pre-existing
        and newly-appended alike) into plain JSON — read from the in-memory
        workbook we already have open, never by re-reading the saved file.
        Call only after save(); the row range is fixed at that point."""
        rows = []
        for row_num in range(FIRST_DATA_ROW, self._next_row):
            values = [
                _serialize_cell(self._sheet.cell(row=row_num, column=col).value)
                for col in range(1, NUM_COLUMNS + 1)
            ]
            rows.append({
                "rowNumber": row_num,
                "values": values,
                "isNew": row_num in self._new_rows,
            })
        return {"columns": list(COLUMNS), "rows": rows}


def _serialize_cell(value):
    """openpyxl can hand back real datetime/date/time objects for cells
    that already held Excel date types (e.g. in a pre-existing custom
    template) — json.dumps can't serialize those directly."""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value
