#!/usr/bin/env python3
"""One-time (re)generation script for templates/Book3.xlsx, the built-in
default workbook. Not part of the request-time pipeline — run this manually
whenever template_schema.py's column layout changes.

Usage:
    python3 scripts/generate_template.py
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from template_schema import (
    COLUMNS,
    FIRST_DATA_ROW,
    NUM_COLUMNS,
    PART_HEADER_ROW,
    SECTION_HEADER_ROW,
)

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "templates", "Book3.xlsx")

COLUMN_WIDTHS = {
    "Product Description": 40,
    "Shipping Bill Date": 14,
    "Invoice Date": 12,
    "Port Code": 10,
    "Shipping Bill No": 12,
}
DEFAULT_WIDTH = 13

PART_FILL = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FIELD_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
PART_FONT = Font(bold=True, size=10)
SECTION_FONT = Font(bold=True, size=9)
FIELD_FONT = Font(bold=True, color="FFFFFF", size=9)
THIN = Side(style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _merge_and_label(sheet, row: int, start_col: int, end_col: int, text: str, fill, font) -> None:
    if end_col > start_col:
        sheet.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.border = BORDER
        cell.fill = fill
        cell.alignment = CENTER
        if font:
            cell.font = font
    sheet.cell(row=row, column=start_col, value=text or None)


def main() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    for start, end, text in PART_HEADER_ROW:
        _merge_and_label(sheet, 1, start, end, text, PART_FILL, PART_FONT)
    # Fill remaining (non-PART, e.g. the blank spacer) columns on row 1 so
    # the whole header band is visually continuous.
    covered = {c for start, end, _ in PART_HEADER_ROW for c in range(start, end + 1)}
    for col in range(1, NUM_COLUMNS + 1):
        if col not in covered:
            _merge_and_label(sheet, 1, col, col, "", PART_FILL, PART_FONT)

    for start, end, text in SECTION_HEADER_ROW:
        _merge_and_label(sheet, 2, start, end, text, SECTION_FILL, SECTION_FONT)
    covered = {c for start, end, _ in SECTION_HEADER_ROW for c in range(start, end + 1)}
    for col in range(1, NUM_COLUMNS + 1):
        if col not in covered:
            _merge_and_label(sheet, 2, col, col, "", SECTION_FILL, SECTION_FONT)

    for col_index, header in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=3, column=col_index, value=header or None)
        cell.font = FIELD_FONT
        cell.fill = FIELD_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        letter = get_column_letter(col_index)
        sheet.column_dimensions[letter].width = COLUMN_WIDTHS.get(header, DEFAULT_WIDTH)

    sheet.row_dimensions[1].height = 20
    sheet.row_dimensions[2].height = 18
    sheet.row_dimensions[3].height = 30
    sheet.freeze_panes = sheet.cell(row=FIRST_DATA_ROW, column=1).coordinate

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    workbook.save(OUTPUT_PATH)
    print(f"Wrote {os.path.abspath(OUTPUT_PATH)}")


if __name__ == "__main__":
    main()
